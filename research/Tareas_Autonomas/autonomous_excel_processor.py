#!/usr/bin/env python3
"""
Autonomous Batch Processor para Excel
Procesa 3905 filas sin preguntar, maneja errores, reporta al final.
BASAO en el patrón Ralph Loop + Checkpointing.
"""

import json
import os
import sys
from datetime import datetime
from typing import Any, Optional

# ============================================
# CONFIGURACIÓN
# ============================================

CHECKPOINT_FILE = ".excel_processor_checkpoint.json"
ERRORS_FILE = "excel_errors.json"
CHUNK_SIZE = 100  # Guardar checkpoint cada N filas

# ============================================
# PROCESSOR CLASS
# ============================================

class AutonomousExcelProcessor:
    """
    Procesa archivos Excel grandes de forma autónoma.
    - Checkpointing: guarda progreso cada CHUNK_SIZE filas
    - Error handling: si falla una fila, continúa
    - Resume: si se corta, reanuda donde quedó
    - Report: resumen final al terminar
    """
    
    def __init__(self, input_file: str, output_file: str, sheet_name: str = 0):
        self.input_file = input_file
        self.output_file = output_file
        self.sheet_name = sheet_name
        self.checkpoint_file = CHECKPOINT_FILE
        self.errors_file = ERRORS_FILE
        self.chunk_size = CHUNK_SIZE
        
        self.processed = 0
        self.errors = []
        self.start_time = datetime.now()
        
    # ---- Checkpoint Management ----
    
    def load_checkpoint(self) -> dict:
        """Carga checkpoint si existe. Si no, empieza desde 0."""
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file) as f:
                    data = json.load(f)
                    print(f"📍 Checkpoint encontrado: fila {data.get('last_row', -1) + 1}")
                    return data
            except (json.JSONDecodeError, IOError):
                pass
        return {'last_row': -1, 'processed': 0, 'timestamp': None}
    
    def save_checkpoint(self, last_row: int):
        """Guarda estado actual."""
        with open(self.checkpoint_file, 'w') as f:
            json.dump({
                'last_row': last_row,
                'processed': self.processed,
                'timestamp': datetime.now().isoformat()
            }, f)
    
    # ---- Error Handling ----
    
    def log_error(self, row_num: int, row_data: Any, error: Exception):
        """Registra error sin detener ejecución."""
        self.errors.append({
            'row': row_num,
            'data_preview': str(row_data)[:100],
            'error_type': type(error).__name__,
            'error': str(error),
            'timestamp': datetime.now().isoformat()
        })
    
    def save_errors(self):
        """Guarda todos los errores al archivo."""
        with open(self.errors_file, 'w', encoding='utf-8') as f:
            json.dump({
                'total_errors': len(self.errors),
                'errors': self.errors
            }, f, indent=2, ensure_ascii=False)
    
    # ---- Processing ----
    
    def process_row(self, row_num: int, row_data: list) -> Optional[dict]:
        """
        PROCESAR UNA FILA.
        
        EDITAR ESTA FUNCIÓN SEGÚN NECESIDADES.
        
        Returns:
            dict con datos procesadas, o None si salta la fila
        Raises:
            Exception si hay error recuperable
        """
        # EJEMPLO - procesar fila Excel:
        # row_data es una lista con los valores de cada celda
        
        # Tu lógica aquí:
        # resultado = transformar(row_data)
        
        # Ejemplo básico - solo devuelve los datos
        return {
            'row': row_num,
            'data': row_data,
            'processed_at': datetime.now().isoformat()
        }
    
    def verify_row(self, result: dict) -> bool:
        """
        VERIFICAR QUE EL RESULTADO ES VÁLIDO.
        
        Editar según criterios de éxito.
        """
        if result is None:
            return False
        # Add verification logic here
        return True
    
    # ---- Main Loop ----
    
    def run(self):
        """Ejecuta procesamiento completo de forma autónoma."""
        
        print("=" * 50)
        print("🚀 AUTONOMOUS EXCEL PROCESSOR")
        print("=" * 50)
        print(f"📁 Input: {self.input_file}")
        print(f"📁 Output: {self.output_file}")
        print(f"📁 Sheet: {self.sheet_name}")
        print()
        
        # Cargar checkpoint
        checkpoint = self.load_checkpoint()
        start_row = checkpoint['last_row'] + 1
        self.processed = checkpoint['processed']
        
        # Intentar importar openpyxl
        try:
            import openpyxl
        except ImportError:
            print("❌ Error: openpyxl no instalado")
            print("   Instala con: pip install openpyxl")
            sys.exit(1)
        
        # Cargar workbook
        print(f"📖 Cargando archivo Excel...")
        try:
            wb = openpyxl.load_workbook(self.input_file, data_only=True)
            ws = wb[self.sheet_name] if isinstance(self.sheet_name, str) else wb.active
            
            # Obtener total de filas
            max_row = ws.max_row
            total_rows = max_row - 1  # Resta 1 por header
            
            print(f"📊 Total filas de datos: {total_rows}")
            print(f"📍 Comenzando desde fila: {start_row + 1}")  # +1 por header
            print()
            
        except Exception as e:
            print(f"❌ Error cargando archivo: {e}")
            sys.exit(1)
        
        # Limpiar archivo de output
        if os.path.exists(self.output_file):
            os.remove(self.output_file)
        
        results = []
        
        print("🔄 Procesando...")
        print("-" * 50)
        
        # Iterar por cada fila (empezando desde checkpoint)
        for row_num in range(start_row + 1, max_row + 1):  # +1 por header row
            try:
                # Leer fila
                row_data = [cell.value for cell in ws[row_num]]
                
                # Procesar
                result = self.process_row(row_num, row_data)
                
                # Verificar
                if result is not None and self.verify_row(result):
                    results.append(result)
                    self.processed += 1
                else:
                    # Row saltada o no válida
                    self.log_error(row_num, row_data, Exception("Verification failed"))
                
            except Exception as e:
                # Error en esta fila - continuar
                self.log_error(row_num, row_data if 'row_data' in locals() else None, e)
                self.processed += 1  # Cuenta como procesada (aunque con error)
            
            # Progress y checkpoint
            if self.processed % self.chunk_size == 0:
                self.save_checkpoint(row_num)
                self.save_errors()
                
                elapsed = (datetime.now() - self.start_time).total_seconds()
                rate = self.processed / elapsed if elapsed > 0 else 0
                remaining = (total_rows - row_num) / rate if rate > 0 else 0
                
                print(f"  ✅ {self.processed}/{total_rows} | Errors: {len(self.errors)} | Rate: {rate:.1f}/s | ETA: {remaining/60:.1f}min")
        
        # Guardar resultados
        print("-" * 50)
        print("💾 Guardando resultados...")
        
        try:
            with open(self.output_file, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"❌ Error guardando output: {e}")
        
        # Guardar checkpoint final y errores
        self.save_checkpoint(max_row - 1)
        self.save_errors()
        
        # Report final
        elapsed = (datetime.now() - self.start_time).total_seconds()
        
        print()
        print("=" * 50)
        print("🎉 PROCESAMIENTO FINALIZADO")
        print("=" * 50)
        print(f"✅ Procesadas exitosamente: {len(results)}")
        print(f"❌ Errors: {len(self.errors)}")
        print(f"⏱️ Tiempo total: {elapsed:.1f}s ({elapsed/60:.1f}min)")
        print(f"📁 Output: {self.output_file}")
        print(f"📁 Errores: {self.errors_file}")
        print(f"📁 Checkpoint: {self.checkpoint_file}")
        print()
        
        if self.errors:
            print("⚠️ Primeros 5 errores:")
            for err in self.errors[:5]:
                print(f"   Fila {err['row']}: {err['error']}")
        
        wb.close()


# ============================================
# MAIN
# ============================================

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python autonomous_excel_processor.py <input.xlsx> <output.json> [sheet_name]")
        print()
        print("Ejemplo:")
        print("  python autonomous_excel_processor.py datos.xlsx resultados.json")
        print("  python autonomous_excel_processor.py datos.xlsx resultados.json Hoja1")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    sheet_name = sys.argv[3] if len(sys.argv) > 3 else 0
    
    if not os.path.exists(input_file):
        print(f"❌ Archivo no encontrado: {input_file}")
        sys.exit(1)
    
    processor = AutonomousExcelProcessor(input_file, output_file, sheet_name)
    processor.run()